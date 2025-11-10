from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.utils.timezone import localtime
from django.core.paginator import Paginator
from django.db.models import Q
from .models import Attendance
from hr.models import Employee
from django.http import HttpResponse
from datetime import datetime, date, time, timedelta
from calendar import monthrange

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# -------------------------------
# Custom Decorators
# -------------------------------

def login_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.session.get('user_authenticated'):
            messages.error(request, 'Please login to access this page.')
            return redirect('login')
        return view_func(request, *args, **kwargs)
    return wrapper


def role_required(allowed_roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            user_role = request.session.get('user_role')
            if not user_role or user_role not in allowed_roles:
                messages.error(request, 'You do not have permission to access this page.')
                return redirect('access_denied')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# -------------------------------
# Attendance Dashboard
# -------------------------------

@login_required
def attendance_dashboard(request):
    user_id = request.session.get('user_id')
    user_role = request.session.get('user_role')
    
    if user_role == 'ADMIN':
        messages.info(request, 'Admins can only view attendance.')
    
    employee = Employee.objects.get(id=user_id)
    today = timezone.now().date()
    today_attendance = Attendance.objects.filter(employee=employee, date=today).first()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'check_in':
            if today_attendance:
                messages.warning(request, 'You have already checked in today.')
            else:
                Attendance.objects.create(
                    employee=employee,
                    date=today,
                    check_in=timezone.now()
                )
                messages.success(request, 'Check-in successful!')
                return redirect('attendance:dashboard')
        
        elif action == 'check_out':
            if not today_attendance:
                messages.error(request, 'You need to check in first.')
            elif today_attendance.check_out:
                messages.warning(request, 'You have already checked out today.')
            else:
                today_attendance.check_out = timezone.now()
                today_attendance.save()
                messages.success(request, 'Check-out successful!')
                return redirect('attendance:dashboard')
                
    
    context = {
        'today_attendance': today_attendance,
        'employee': employee,
    }
    return render(request, 'attendance/dashboard.html', context)


# -------------------------------
# View All Attendance (Employee)
# -------------------------------

@login_required
def all_attendance(request):
    user_id = request.session.get('user_id')
    user_role = request.session.get('user_role')
    
   
    employee = Employee.objects.get(id=user_id)
    attendance_records = Attendance.objects.filter(employee=employee).order_by('-date')

    # ✅ Month filter (YYYY-MM)
    month_filter = request.GET.get('month', '')
    today = date.today()

    if month_filter:
        year, month = map(int, month_filter.split('-'))
        start_date = date(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = date(year, month, last_day)

        # ✅ Limit end date to today (no future dates)
        if year == today.year and month == today.month:
            end_date = today
        elif date(year, month, last_day) > today:
            end_date = today
    else:
        end_date = today
        start_date = end_date - timedelta(days=30)

    # ✅ Generate list of valid working dates (excluding Sundays)
    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
        if (start_date + timedelta(days=i)).weekday() != 6
        and (start_date + timedelta(days=i)) <= today
    ]

    attendance_dict = {att.date: att for att in attendance_records}
    full_attendance_list = []

    for d in reversed(all_dates):
        record = attendance_dict.get(d)

        if record:
            record.duration_display = "-"
            record.extra_hours_display = "-"

            if record.check_in and record.check_out:
                diff = record.check_out - record.check_in
                total_minutes = diff.total_seconds() / 60
                hours = int(total_minutes // 60)
                minutes = int(total_minutes % 60)
                record.duration_display = f"{hours}h {minutes}m" if hours or minutes else "0 minutes"

                weekday = record.date.weekday()
                if weekday == 5:
                    standard_hours = 6 if employee.location.strip().lower() == 'bbsr' else 4
                else:
                    standard_hours = 9

                extra_hours = (total_minutes / 60) - standard_hours
                if extra_hours > 0:
                    eh_hours = int(extra_hours)
                    eh_minutes = int((extra_hours - eh_hours) * 60)
                    record.extra_hours_display = f"{eh_hours}h {eh_minutes}m" if eh_hours or eh_minutes else "0h 0m"
                else:
                    record.extra_hours_display = "0h 0m"
            elif record.check_in and not record.check_out:
                record.duration_display = "In Progress"
                record.extra_hours_display = "-"
            else:
                record.duration_display = "-"
                record.extra_hours_display = "-"
            full_attendance_list.append(record)
        else:
            fake_record = Attendance(
                employee=employee,
                date=d,
                check_in=None,
                check_out=None,
            )
            fake_record.duration_display = "-"
            fake_record.extra_hours_display = "-"
            full_attendance_list.append(fake_record)

    # ✅ Sort latest first (no pagination)
    full_attendance_list.sort(key=lambda x: x.date, reverse=True)

      # ✅ Determine selected month (default: current)
    if not month_filter:
        month_filter = today.strftime("%Y-%m")

    context = {
        'attendances': full_attendance_list,  # ✅ Full list shown in one page
        'employee': employee,
        'selected_month': month_filter, # ✅ Always valid (e.g. 2025-10)
        'today': today,  
    }
    return render(request, 'attendance/all_attendance.html', context)

# -------------------------------
# Admin / HR Attendance Report
# -------------------------------

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def attendance_report(request):
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    branch = request.GET.get('branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status_filter', '')

    today = date.today()
    office_start_time = time(9, 30)

    # ✅ Parse date range safely (default to today if blank)
    if date_from:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        except ValueError:
            start_date = today
    else:
        start_date = today

    if date_to:
        try:
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except ValueError:
            end_date = today
    else:
        end_date = today

    # ✅ Prevent selecting a range beyond today or inverted
    if end_date > today:
        end_date = today
    if end_date < start_date:
        end_date = start_date

    # ✅ Filter employees
    employees = Employee.objects.all().order_by('first_name', 'last_name')

    if department:
        employees = employees.filter(department=department)
    if hasattr(Employee, 'branch') and branch:
        employees = employees.filter(branch=branch)
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(employee_id__icontains=search_query)
        )

    # ✅ Get all dates in the selected range
    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    attendance_data = []

    # ✅ Build report for every employee and every date
    for emp in employees:
        for day in all_dates:
            att = Attendance.objects.filter(employee=emp, date=day).first()

            record = {
                'employee_pk': emp.id,
                'employee_id': emp.employee_id,
                'employee_name': f"{emp.first_name} {emp.last_name}",
                'department': emp.department,
                'branch': getattr(emp, 'branch', '—'),
                'date': day,
                'check_in': "—",
                'check_out': "—",
                'status': 'Absent',
                'punctuality': '-',
                'duration_display': '-'
            }

            if att:
                if att.check_in:
                    record['check_in'] = localtime(att.check_in).strftime("%I:%M %p")
                if att.check_out:
                    record['check_out'] = localtime(att.check_out).strftime("%I:%M %p")

                # ✅ Status logic
                if att.check_in and att.check_out:
                    record['status'] = "Present"
                elif att.check_in and not att.check_out:
                    record['status'] = "Half Day"

                # ✅ Punctuality logic
                if att.check_in:
                    check_in_time = localtime(att.check_in).time()
                    record['punctuality'] = "On Time" if check_in_time <= office_start_time else "Late"

                # ✅ Duration logic
                if att.check_in and att.check_out:
                    diff = att.check_out - att.check_in
                    total_minutes = diff.total_seconds() / 60
                    hours = int(total_minutes // 60)
                    minutes = int(total_minutes % 60)
                    record['duration_display'] = f"{hours}h {minutes}m"
                elif att.check_in and not att.check_out:
                    record['duration_display'] = "In Progress"

            attendance_data.append(record)

    # ✅ Apply status filter
    if status_filter:
        if status_filter in ['Present', 'Absent', 'Half Day']:
            attendance_data = [a for a in attendance_data if a['status'] == status_filter]
        elif status_filter in ['On Time', 'Late']:
            attendance_data = [a for a in attendance_data if a['punctuality'] == status_filter]

    # ✅ Sort employees alphabetically by name (A → Z)
    attendance_data.sort(key=lambda x: x['employee_name'].lower())

    # ✅ Paginate results (20 per page)
    paginator = Paginator(attendance_data, 20)
    page = request.GET.get('page')
    attendance_records = paginator.get_page(page)

    # ✅ Dropdown data
    departments = Employee.objects.values_list('department', flat=True).distinct()
    branches = (
        Employee.objects.values_list('branch', flat=True).distinct()
        if hasattr(Employee, 'branch')
        else ['Bhubaneswar', 'Bangalore', 'Jaipur']
    )

    context = {
        'attendances': attendance_records,
        'departments': departments,
        'branches': branches,
        'search_query': search_query,
        'selected_department': department,
        'selected_branch': branch,
        'date_from': date_from,
        'date_to': date_to,
        'today': today,
        'status_filter': status_filter,
    }

    return render(request, 'attendance/report.html', context)

@login_required
@role_required(['ADMIN', 'HR', 'SUPER ADMIN'])
def download_admin_attendance_report(request):
    """Download Excel with exactly the same filtered data shown in the dashboard."""
    search_query = request.GET.get('search', '')
    department = request.GET.get('department', '')
    branch = request.GET.get('branch', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status_filter = request.GET.get('status_filter', '')

    today = date.today()
    office_start_time = time(9, 30)

    # ✅ Parse exact date range — only use today if both filters are empty
    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
    except ValueError:
        start_date = None

    try:
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except ValueError:
        end_date = None

    # If both filters missing → default to today
    if not start_date and not end_date:
        start_date = end_date = today
    elif start_date and not end_date:
        end_date = start_date
    elif end_date and not start_date:
        start_date = end_date

    # Prevent selecting a future date
    if end_date > today:
        end_date = today

    # ✅ Get filtered employees (same as dashboard)
    employees = Employee.objects.all().order_by('first_name', 'last_name')

    if department:
        employees = employees.filter(department=department)
    if hasattr(Employee, 'branch') and branch:
        employees = employees.filter(branch=branch)
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(employee_id__icontains=search_query)
        )

    # ✅ Get all dates within range
    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
    ]

    attendance_data = []

    # ✅ Build records for each employee for selected date(s)
    for emp in employees:
        for day in all_dates:
            att = Attendance.objects.filter(employee=emp, date=day).first()

            record = {
                'employee_id': emp.employee_id,
                'employee_name': f"{emp.first_name} {emp.last_name}",
                'department': emp.department,
                'branch': getattr(emp, 'branch', '—'),
                'date': day.strftime("%b %d, %Y"),
                'check_in': "—",
                'check_out': "—",
                'status': 'Absent',
                'punctuality': '-',
                'duration_display': '-'
            }

            if att:
                if att.check_in:
                    record['check_in'] = localtime(att.check_in).strftime("%I:%M %p")
                if att.check_out:
                    record['check_out'] = localtime(att.check_out).strftime("%I:%M %p")

                # Status logic
                if att.check_in and att.check_out:
                    record['status'] = "Present"
                elif att.check_in and not att.check_out:
                    record['status'] = "Half Day"

                # Punctuality
                if att.check_in:
                    check_in_time = localtime(att.check_in).time()
                    record['punctuality'] = "On Time" if check_in_time <= office_start_time else "Late"

                # Duration
                if att.check_in and att.check_out:
                    diff = att.check_out - att.check_in
                    total_minutes = diff.total_seconds() / 60
                    hours = int(total_minutes // 60)
                    minutes = int(total_minutes % 60)
                    record['duration_display'] = f"{hours}h {minutes}m"
                elif att.check_in and not att.check_out:
                    record['duration_display'] = "In Progress"

            attendance_data.append(record)

    # ✅ Apply filters just like dashboard
    if status_filter:
        if status_filter in ['Present', 'Absent', 'Half Day']:
            attendance_data = [a for a in attendance_data if a['status'] == status_filter]
        elif status_filter in ['On Time', 'Late']:
            attendance_data = [a for a in attendance_data if a['punctuality'] == status_filter]

    # ✅ Sort alphabetically
    attendance_data.sort(key=lambda x: x['employee_name'].lower())

    # ✅ If no attendance data (filtered range has none)
    if not attendance_data:
        response = HttpResponse(
            "No attendance data found for the selected date range and filters.",
            content_type="text/plain"
        )
        response['Content-Disposition'] = 'attachment; filename=\"Empty_Attendance_Report.txt\"'
        return response

    # ✅ Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    headers = [
        "Employee ID", "Employee Name", "Department", "Branch",
        "Date", "Check-In", "Check-Out", "Status", "Punctuality", "Duration"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style

    for record in attendance_data:
        ws.append([
            record['employee_id'],
            record['employee_name'],
            record['department'],
            record['branch'],
            record['date'],
            record['check_in'],
            record['check_out'],
            record['status'],
            record['punctuality'],
            record['duration_display'],
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = align_center
            cell.border = border_style

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    filename = f"Attendance_Report_{start_date.strftime('%b_%d_%Y')}_to_{end_date.strftime('%b_%d_%Y')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=\"{filename}\"'
    wb.save(response)
    return response



# -------------------------------
# Generate Attendance Report PDF
# -------------------------------

@login_required
def download_attendance_report_excel(request):
    """Download Excel attendance report for a single employee (monthly view)."""
    user_id = request.session.get('user_id')
    employee = Employee.objects.get(id=user_id)

    # ✅ Get selected month
    month_filter = request.GET.get('month', '')
    today = date.today()

    if month_filter:
        year, month = map(int, month_filter.split('-'))
    else:
        year, month = today.year, today.month

    # ✅ Month range
    start_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    # ✅ Restrict to current day for live month
    if year == today.year and month == today.month:
        end_date = today
    elif end_date > today:
        end_date = today

    # ✅ Get existing attendance data
    attendance_records = Attendance.objects.filter(
        employee=employee,
        date__range=[start_date, end_date]
    )
    attendance_dict = {a.date: a for a in attendance_records}

    # ✅ Generate list of all working days (excluding Sundays)
    all_dates = [
        start_date + timedelta(days=i)
        for i in range((end_date - start_date).days + 1)
        if (start_date + timedelta(days=i)).weekday() != 6  # Exclude Sundays
    ]

    # ✅ Prepare full attendance list (with Absent days)
    full_attendance_list = []
    for d in all_dates:
        if d in attendance_dict:
            record = attendance_dict[d]
            check_in = localtime(record.check_in).strftime("%I:%M %p") if record.check_in else "-"
            check_out = localtime(record.check_out).strftime("%I:%M %p") if record.check_out else "-"
            if record.check_in and record.check_out:
                status = "Present"
                diff = record.check_out - record.check_in
                total_minutes = diff.total_seconds() / 60
                hours = int(total_minutes // 60)
                minutes = int(total_minutes % 60)
                duration = f"{hours}h {minutes}m"
            elif record.check_in:
                status = "Half Day"
                duration = "In Progress"
            else:
                status = "Absent"
                duration = "-"
        else:
            check_in = "-"
            check_out = "-"
            status = "Absent"
            duration = "-"
        full_attendance_list.append({
            'date': d.strftime("%b %d, %Y"),
            'check_in': check_in,
            'check_out': check_out,
            'status': status,
            'duration': duration,
        })

    # ✅ Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    align_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ✅ Header
    headers = ["Date", "Check-In", "Check-Out", "Status", "Duration"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_style

    # ✅ Data Rows
    for record in full_attendance_list:
        ws.append([
            record['date'],
            record['check_in'],
            record['check_out'],
            record['status'],
            record['duration']
        ])

    # ✅ Format Cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = align_center
            cell.border = border_style

    # ✅ Auto column widths
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    # ✅ File Name
    month_name = start_date.strftime("%B")
    filename = f"{employee.first_name}_{month_name}_{year}_Attendance_Report.xlsx"

    # ✅ Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response